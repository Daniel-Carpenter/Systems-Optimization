"""
@author: Daniel Carpenter

Knapsack Function with Multiple Bags
Given a number of bags and weight capcity of each bag, what is the optimal 
way to pack these bags?
"""

# Import Gurobi
from gurobipy import *

# Import excel reading package
import openpyxl as opxl


# Inputs =====================================================================

## The bags you want to pack
Bags = ['Portfolio']

## Capacity of each bags
bagCapacities = {'Portfolio':200}

## Names of the Excel file and sheet
xlFileName  = 'ConstructionProjects'
xlSheetName = 'ConstructionProjects'

# ============================================================================

# Create the model
m = Model('General Knapsack - Packing Multiple Bags given weight constraint')
m.setParam(GRB.Param.OutputFlag, 0)

# SETS =======================================================================

# The item and its number that you are considering packing
Items = []

# PARAMETERS =================================================================

## weight (in pounds) of each item that you are considering packing
weight = {}

## Intrinsic value of each item that you are considering packing
value = {}

## Read in the values of the Items, weight, and value in the 'Data.xlxs' file

xlFile = opxl.load_workbook(xlFileName + ".xlsx")

### First Row to read in minus 1
row = 2

## Loop through each row and col and insert data into sets and parameters
while xlFile[xlSheetName].cell(row = row + 1, column = 1).value:
    
    #### Read in item name
    newItem = xlFile[xlSheetName].cell(row = row + 1, column = 1).value
    Items.append(newItem)
    
    #### Read in weight col
    weight[newItem] = xlFile[xlSheetName].cell(row = row + 1, column = 2).value
    
    #### read in value col
    value[newItem] = xlFile[xlSheetName].cell(row = row + 1, column = 3).value
    
    #### Increment row
    row += 1


# DESCISION VARIABLES ========================================================
isPacked = m.addVars(Bags, Items, vtype = GRB.BINARY)


# OPTIMIZATION ===============================================================

## Add the new constraint for two bags capacity
m.addConstrs(quicksum(weight[item] * isPacked[bag, item] for item in Items) 
                     <= bagCapacities[bag] for bag in Bags)

## C2: The item cannot be packed in multiple bags
PACKED = 1

### The sum of the binary var `item` (over all bags '*') cannot exceed a count of 1
for item in Items:
        m.addConstr(isPacked.sum('*', item) <= PACKED)

## Maximize the intrinsic value of the packed items for each bag
objFun = quicksum(quicksum(value[item] * isPacked[bag, item] 
                           for item in Items) 
                  for bag in Bags)

## Set objective, update the model, and Optimize
m.setObjective(objFun, GRB.MAXIMIZE)
m.update()
m.optimize()

# Create print function
def printOptimalSolution(model, isPacked, weight, value, Items, Bags):
    if model.status == GRB.Status.OPTIMAL:
        print('Optimizing your vacation: maximize value of packed items while not packing too much')
        
        # Print the optimal intrinsic value
        print("\n*Optimal value of packed items: " 
              + "{:,.4f}".format(model.objVal) + " total intrinsic value")
        
        for bag in Bags:
            
            print('\n\n======== ' + bag + " ========")
        
            # GEt the total weight of the bag
            totalWeight = sum(weight[item] * isPacked[bag, item].x 
                              for item in Items)
            
            # print the total amount packed in lbs
            print('\nTotal Weight of bag: ' + str(totalWeight) + ' lbs.')
            
            # Print the items packed with weight and value
            print("\nItems that you should pack:")
            print("Item | isPacked \t| lbs. \t| Intrinsic Value")
            for item in Items:
                if isPacked[bag, item].x > 0.5:
                    print(str(item) + ": " 
                          + "\t"    + str(isPacked[bag, item].x)
                          + "\t |\t" + str(  weight[item])
                          + "\t |\t" + str(   value[item]))
        
            # Print the items NOT packed with weight and value
            print("\nItems that you should NOT pack:")
            print("Item | isPacked \t| lbs. \t| Intrinsic Value")
            for item in Items:
                if isPacked[bag, item].x < 0.5:
                    print(str(item) + ": " 
                          + "\t "    + str(isPacked[bag, item].x)
                          + "\t|\t" + str(  weight[item])
                          + "\t|\t" + str(   value[item]))
    
print('\nOptimization with multiple bags')
printOptimalSolution(m, isPacked, weight, value, Items, Bags)

