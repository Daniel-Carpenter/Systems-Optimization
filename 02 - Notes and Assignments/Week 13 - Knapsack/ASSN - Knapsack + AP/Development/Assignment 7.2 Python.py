"""
Assignment 7
@author: Daniel.Carpenter
"""

from gurobipy import *
import openpyxl as opxl

"""
Problem 2.1
"""


# EXCEL READ ==================================================================

# Excel file and sheet name
fileXLS  = "OptiCoffee.xlsx"
sheetXLS = "Information"

doc = opxl.load_workbook(fileXLS)


# Read in single value variables ----------------------------------------------

## First row num less one
rowNum = 9 - 1

## s: Marginal costs of stockout backlog [thousands of USD/Ton]
s = doc[sheetXLS].cell(row = rowNum + 1, column = 2).value

## h: Hiring and training costs [thousands of USD/worker]
h = doc[sheetXLS].cell(row = rowNum + 2, column = 2).value

## f: Layoff costs [thousands of USD/worker]
f = doc[sheetXLS].cell(row = rowNum + 3, column = 2).value

## k: Labour hours required per ton produced [hrs/Ton]
k = doc[sheetXLS].cell(row = rowNum + 4, column = 2).value

## w: Regular time cost [thousands of UDS/worker/hr]
w = doc[sheetXLS].cell(row = rowNum + 5, column = 2).value

## o: Overtime cost [thousands of UDS/hr]
o = doc[sheetXLS].cell(row = rowNum + 6, column = 2).value

## c: Subcontarcting costs [thousands of UDS/Ton]
c = doc[sheetXLS].cell(row = rowNum + 7, column = 2).value

## a: Initial inventory [Ton]
a = doc[sheetXLS].cell(row = rowNum + 8, column = 2).value

## b: Initial workforce [worker]
b = doc[sheetXLS].cell(row = rowNum + 9, column = 2).value

## e: Initial backlog [worker]
e = doc[sheetXLS].cell(row = rowNum + 10, column = 2).value


# Read in Multivalued Monthly data --------------------------------------------

## First row num less one
rowNum = 3 - 1
firstCol = 1

## Number of periods in the sample
numPeriods = 12

## T: Set of months or periods
T = [doc[sheetXLS].cell(row = rowNum + 1, column = firstCol + col).value for col in range(1, numPeriods + 1)]

## d: Demand [Ton]
d = {T[col - 1]:doc[sheetXLS].cell(row = rowNum + 2, column = firstCol + col).value for col in range(1, numPeriods + 1)}

## p: Unit production costs [thousands of USD/Ton]
p = {T[col - 1]:doc[sheetXLS].cell(row = rowNum + 3, column = firstCol + col).value for col in range(1, numPeriods + 1)}

## i: Inventory holding costs [thousands of USD/Ton]
i = {T[col - 1]:doc[sheetXLS].cell(row = rowNum + 4, column = firstCol + col).value for col in range(1, numPeriods + 1)}

## n: Number of regular working hours [hrs]
n = {T[col - 1]:doc[sheetXLS].cell(row = rowNum + 5, column = firstCol + col).value for col in range(1, numPeriods + 1)}

## m: Maximum number of overtime hours per worker [hrs/worker]
m = {T[col - 1]:doc[sheetXLS].cell(row = rowNum + 6, column = firstCol + col).value for col in range(1, numPeriods + 1)}


# OPTIMIZATION MODEL ==========================================================

## Create the model -----------------------------------------------------------
m2 = Model('OptiCoffee')
m2.setParam(GRB.Param.OutputFlag, 0)


## Decision Variables ---------------------------------------------------------

W = {t:m2.addVar(name = "W["+str(t)+"]", vtype = GRB.INTEGER)    for t in T}
O = {t:m2.addVar(name = "O["+str(t)+"]", vtype = GRB.CONTINUOUS) for t in T}
H = {t:m2.addVar(name = "H["+str(t)+"]", vtype = GRB.INTEGER)    for t in T}
F = {t:m2.addVar(name = "F["+str(t)+"]", vtype = GRB.INTEGER)    for t in T}
P = {t:m2.addVar(name = "P["+str(t)+"]", vtype = GRB.CONTINUOUS) for t in T}
I = {t:m2.addVar(name = "I["+str(t)+"]", vtype = GRB.CONTINUOUS) for t in T}
S = {t:m2.addVar(name = "S["+str(t)+"]", vtype = GRB.CONTINUOUS) for t in T}
C = {t:m2.addVar(name = "C["+str(t)+"]", vtype = GRB.CONTINUOUS) for t in T}


## Objective Function ---------------------------------------------------------

### Regular Time Labor Cost
RTLC = quicksum(n[t] * w * W[t] for t in T)

### Overtime Labor Cost
OTLC = quicksum(o * O[t]        for t in T)

### Cost of Hiring
HC   = quicksum(h * H[t]        for t in T)

### Cost of Layoffs
FC   = quicksum(f * F[t]        for t in T)

### Cost of Holding Inventory
HIC  = quicksum(i[t] * I[t]     for t in T)

### Cost of Stocking Out
CSO  = quicksum(s * S[t]        for t in T)

### Production Costs
PC   = quicksum(p[t] * P[t]     for t in T)

### Sub-contracting cost
SC   = quicksum(c * C[t]        for t in T)

### Initilize the Actual Objective Function
FO = (RTLC + OTLC + HC + FC + HIC + CSO + PC + SC)

### Set the Objective Function to m2 Model
m2.setObjective(FO, GRB.MINIMIZE)


# CONSTRAINTS =================================================================

## Workforce, hiring, and layoffs
m2.addConstr (W[1] == b + H[1] - F[1])
m2.addConstrs(W[t] == W[t - 1] + H[t] - F[t] for t in T if t is not T[0])

## Capacity Constraints
m2.addConstrs(k * P[t] <= n[t] * W[t] + O[t] for t in T)

## Inventory Balance Constraints
m2.addConstr (a + P[1] + C[1] - e               == d[1] + I[1] - S[1])
m2.addConstrs(I[t - 1] + P[t] + C[t] - S[t - 1] == d[t] + I[t] - S[t] 
              for t in T if t is not T[0])

## Overtime Constraints
m2.addConstrs(O[t] <= m[t] * W[t] for t in T)

## Additional Constraints -----------------------------------------------------

FIRST_MONTH = 1

### Initial Inventory
m2.addConstr(I[FIRST_MONTH] == a)

### Initial Workforce
m2.addConstr(W[FIRST_MONTH] == b)

### Initial Backlog
m2.addConstr(S[FIRST_MONTH] == e)


## Optimize the model ---------------------------------------------------------
m2.update()
m2.optimize()


# PRINT SOLUTION ==============================================================

## Print to Console -----------------------------------------------------------
if m2.status == GRB.OPTIMAL:
    print('\n============= PROBLEM 2.1 =============')
    print('\nOptimal/Minimized Cost: %4f' % m2.objVal)
    print('\nDecision Variables ----------------------')
    for var in m2.getVars():
        print('%s\t%g' % (var.varName, var.x))


## Write to Excel -------------------------------------------------------------

import xlwt
from xlwt import Workbook

### Create the workbook
wb = Workbook()

### Add the sheet name
s1 = wb.add_sheet('OptiCoffeeOutput')

### Write the column headers
s1.write(0, 0, 'Period')
s1.write(0, 1, 'Demand')
s1.write(0, 2, 'Production')
s1.write(0, 3, 'SubContracting')
s1.write(0, 4, 'Backlogs')
s1.write(0, 5, 'Inventory')
s1.write(0, 6, 'Workforce')
s1.write(0, 7, 'Hirings')
s1.write(0, 8, 'Layoffs')
s1.write(0, 9, 'OvertimeHours')

### Print the Values
row = 1
for t in T:
    s1.write(row, 0, t)
    s1.write(row, 1, d[t])
    s1.write(row, 2, P[t].x)
    s1.write(row, 3, C[t].x)
    s1.write(row, 4, S[t].x)
    s1.write(row, 5, I[t].x)
    s1.write(row, 6, W[t].x)
    s1.write(row, 7, H[t].x)
    s1.write(row, 8, F[t].x)
    s1.write(row, 9, O[t].x)
    row += 1

### Save the file
wb.save("OptiCoffee Analysis Output.xls")


"""
Problem 2.2
"""

## Only 4 production periods
periodLimit = 4

## Count if production is greater than 0 in a given period, store in binary list.
## sum of the output binary numbers must be less or equal to than the limit (4)
m2.addConstr(sum(P[t].x > 0 for t in T) <= periodLimit )

## Optimize the model ---------------------------------------------------------
m2.update()
m2.optimize()


# PRINT SOLUTION ==============================================================

## Print to Console -----------------------------------------------------------
if m2.status == GRB.OPTIMAL:
    print('\n============= PROBLEM 2.2 =============')
    print('\nOptimal/Minimized Cost: %4f' % m2.objVal)
    print('\nDecision Variables ----------------------')
    for var in m2.getVars():
        print('%s\t%g' % (var.varName, var.x))


## Write to Excel -------------------------------------------------------------

import xlwt
from xlwt import Workbook

### Create the workbook
wb = Workbook()

### Add the sheet name
s1 = wb.add_sheet('OptiCoffeeOutput')

### Write the column headers
s1.write(0, 0, 'Period')
s1.write(0, 1, 'Demand')
s1.write(0, 2, 'Production')
s1.write(0, 3, 'SubContracting')
s1.write(0, 4, 'Backlogs')
s1.write(0, 5, 'Inventory')
s1.write(0, 6, 'Workforce')
s1.write(0, 7, 'Hirings')
s1.write(0, 8, 'Layoffs')
s1.write(0, 9, 'OvertimeHours')

### Print the Values
row = 1
for t in T:
    s1.write(row, 0, t)
    s1.write(row, 1, d[t])
    s1.write(row, 2, P[t].x)
    s1.write(row, 3, C[t].x)
    s1.write(row, 4, S[t].x)
    s1.write(row, 5, I[t].x)
    s1.write(row, 6, W[t].x)
    s1.write(row, 7, H[t].x)
    s1.write(row, 8, F[t].x)
    s1.write(row, 9, O[t].x)
    row += 1

### Save the file
wb.save("OptiCoffee Analysis Output Period Limit (2.2).xls")


"""
This problem would run and if the new constraint worked properly.
Essentially, by goal was to add a constraint with the following logic:
    Ensure that the count of production periods are less than the 
    production period limit of 4. However, it seems that this count
    operation does not work well with Gurobi's algorithm.

"""





























