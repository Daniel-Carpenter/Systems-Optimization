"""
Assignment 7
@author: Daniel.Carpenter
"""

from gurobipy import *
import openpyxl as opxl

# SETS & PARAMETERS ==========================================================
Projects      = []
cost          = {}
jobsGenerated = {}


# Read data from Excel -------------------------------------------------------

## Make connection with the file and store in object
xlFile = opxl.load_workbook("ConstructionProjects.xlsx")
row = 1

## Loop through each row and col and insert data into sets and parameters
while xlFile["ConstructionProjects"].cell(row = row + 1, column = 2).value:
    
    ### get the project number and add to the set of Projects
    project = xlFile["ConstructionProjects"].cell(row = row + 1, column = 1).value
    Projects.append(project)
    
    ### Get the cost of the individual project and store data
    cost[project]          = xlFile["ConstructionProjects"].cell(row = row + 1, column = 2).value
    
    ### Get the number of jobs generated for the individual project and store data
    jobsGenerated[project] = xlFile["ConstructionProjects"].cell(row = row + 1, column = 3).value
    row += 1

## Validate correct Reading of Excel Data    
print("\nThe Projects")
print(Projects)

print("\nThe Costs")
print(cost)

print("\nThe Jobs Generated")
print(jobsGenerated)


## The budget for all projects: millions of USD 
budget = 200

# OPTIMIZATION MODEL =========================================================

## Model
m1 = Model('Construction Projects Optimization')
m1.setParam(GRB.Param.OutputFlag, 0)

## Variables 
isExecuted = {project : m1.addVar(vtype = GRB.BINARY, 
                                 name = "isExecuted[" + str(project) + "]") 
              for project in Projects}

## Constraints
m1.addConstr(quicksum(cost[project] * isExecuted[project] for project in Projects) <= budget)

## Objective Function
objFun = quicksum(jobsGenerated[project] * isExecuted[project] for project in Projects)

## Optimize the model and set the model sense to maximize
m1.setObjective(objFun, GRB.MAXIMIZE)
m1.update()
m1.optimize()

# PRINT RESULTS ==============================================================

# Print the solution to the console
print("MODEL 1 (Original")
if m1.status == GRB.Status.OPTIMAL:
    print("\n*Optimal number of jobs that the project generated: " 
          + "{:,.4f}".format(m1.objVal) + " thousand jobs")
    print('\n**Below shows each project that was executed, their cost, and the jobs generated.')
    print('  Note that if a project in the set of Projects does not exist here,')
    print('  Then including it does not given an optimial solution. Therefore ignored.\n')
    print("\tProject\t| Executed\t| Cost\t| Jobs Generated")
    for project in Projects:
        if isExecuted[project].x > 0.5:
            print("\tproject[" + str(project) + "] =" 
                  + "\t"     + str(   isExecuted[project].x)
                  + "\t|\t" + str(         cost[project])
                  + "\t|\t" + str(jobsGenerated[project]))



# PROBLEM 1 (c) NEW CONDITION ================================================

## New Constraints 

### Variables constain the effected project numbers
CONDITIONAL_PROJECT_8 = 8
CONDITIONAL_PROJECT_9 = 9
SUBJECTIVE_PROJECT_10 = 10
PROJECT_IS_EXECUTED   = 1

### Requirement to execute construction project #10 if construction projects #8 and #9 are executed 
if  isExecuted[CONDITIONAL_PROJECT_8].x > 0.5 and isExecuted[CONDITIONAL_PROJECT_9].x > 0.5:
    m1.addConstr(isExecuted[SUBJECTIVE_PROJECT_10] == PROJECT_IS_EXECUTED)

### Reoptimize model
m1.update()
m1.optimize()

## PRINT THE MODEL
print("\n---------------------------------------------------------------------------------")
print("Model with requirement that project 10 is executed if project 8 and 9 are executed")
if m1.status == GRB.Status.OPTIMAL:
    print("\n*Optimal number of jobs that the project generated: " 
          + "{:,.4f}".format(m1.objVal) + " thousand jobs")
    print('\n**Below shows each project that was executed, their cost, and the jobs generated.')
    print('  Note that if a project in the set of Projects does not exist here,')
    print('  Then including it does not given an optimial solution. Therefore ignored.\n')
    print("\tProject\t| Executed\t| Cost\t| Jobs Generated")
    for project in Projects:
        if isExecuted[project].x > 0.5:
            print("\tproject[" + str(project) + "] =" 
                  + "\t"     + str(   isExecuted[project].x)
                  + "\t|\t" + str(         cost[project])
                  + "\t|\t" + str(jobsGenerated[project]))
            
    print(                  
    """
    Clearly, you can tell that project 10 now is executed but the value of the optimal
    solution has stayed the same.
    """
    )
    



