"""
Knapsack Example
"""

"""
PART I:
    Imagine that you are planning your next vacations.
    For your trip, first you want to determine which items to pack. 
    To take this decision, you make a list of 30 items from which you want to 
    pick the ones to be packed. For each of these 30 items, you know 
    its weight (in pounds) and its associated intrinsic value 
    (the more you want to take the item, the higher its “value”)
"""

# Import Gurobi
from gurobipy import *

# Import excel reading package
import openpyxl as opxl

# Create the model
m = Model('Knapsack - Packing for Vacation')
m.setParam(GRB.Param.OutputFlag, 0)


# SETS =======================================================================

# The item and its number that you are considering packing
Items = []
Bags  = ['Bag 1']

# PARAMETERS =================================================================

## weight (in pounds) of each item that you are considering packing
weight = {}

## Intrinsic value of each item that you are considering packing
value = {}

## Read in the values of the Items, weight, and value in the 'Data.xlxs' file
xlFileName  = 'Data'
xlSheetName = 'Data'

xlFile = opxl.load_workbook(xlFileName + ".xlsx")

### First Row to read in minus 1
row = 5 - 1

## Loop through each row and col and insert data into sets and parameters
while xlFile[xlSheetName].cell(row = row + 1, column = 2).value:
    
    #### Read in item name
    newItem = xlFile[xlSheetName].cell(row = row + 1, column = 2).value
    Items.append(newItem)
    
    #### Read in weight col
    weight[newItem] = xlFile[xlSheetName].cell(row = row + 1, column = 3).value
    
    #### read in value col
    value[newItem] = xlFile[xlSheetName].cell(row = row + 1, column = 4).value
    
    #### Increment row
    row += 1


## Capacity of one bag = 15 lbs
bagCapacities = {'Bag 1':15}

# DESCISION VARIABLES ========================================================
isPacked = m.addVars(Bags, Items, vtype = GRB.BINARY)


# OPTIMIZATION ===============================================================

## Add the new constraint for two bags capacity
m.addConstrs(quicksum(weight[item] * isPacked[bag, item] for item in Items) 
                     <= bagCapacities[bag] for bag in Bags)

## Maximize the intrinsic value of the packed items for each bag
objFun = quicksum(quicksum(value[item] * isPacked[bag, item] 
                           for item in Items) 
                  for bag in Bags)

## Set objective, update the model, and Optimize
m.setObjective(objFun, GRB.MAXIMIZE)
m.update()
m.optimize()


# OUTPUT =====================================================================

# Create print function
def printOptimalSolution(model, isPacked, weight, value, Items, Bags):
    if model.status == GRB.Status.OPTIMAL:
        print('Optimizing your vacation: maximize value of packed items while not packing too much')
        
        # Print the optimal intrinsic value
        print("\n*Optimal value of packed items: " 
              + "{:,.4f}".format(model.objVal) + " total intrinsic value")
        
        for bag in Bags:
            
            print('\n\n======== ' + bag + " ========")
        
            # GEt the total weight of the bag
            totalWeight = sum(weight[item] * isPacked[bag, item].x 
                              for item in Items)
            
            # print the total amount packed in lbs
            print('\nTotal Weight of bag: ' + str(totalWeight) + ' lbs.')
            
            # Print the items packed with weight and value
            print("\nItems that you should pack:")
            print("Item | isPacked \t| lbs. \t| Intrinsic Value")
            for item in Items:
                if isPacked[bag, item].x > 0.5:
                    print(str(item) + ": " 
                          + "\t"    + str(isPacked[bag, item].x)
                          + "\t |\t" + str(  weight[item])
                          + "\t |\t" + str(   value[item]))
        
            # Print the items NOT packed with weight and value
            print("\nItems that you should NOT pack:")
            print("Item | isPacked \t| lbs. \t| Intrinsic Value")
            for item in Items:
                if isPacked[bag, item].x < 0.5:
                    print(str(item) + ": " 
                          + "\t "    + str(isPacked[bag, item].x)
                          + "\t|\t" + str(  weight[item])
                          + "\t|\t" + str(   value[item]))
    
            
# Print the solution to the console
print('\nPART I: ONLY ONE BAG')
printOptimalSolution(m, isPacked, weight, value, Items, Bags)
            
"""
PART II:
    Now, assume that for your trip you can use two bags. 
    The first  bag has a maximum capacity of 15 pounds and 
    the second one has a maximum capacity of 50 pounds.
"""

# SETS =======================================================================
# New SET defining the bags
Bags.append('Bag 2')

# Second Bag Capacity = 50 lbs
bagCapacities = {'Bag 1':15, 'Bag 2':50}


# DESCISION VARIABLES ========================================================
isPacked = m.addVars(Bags, Items, vtype = GRB.BINARY)


# OPTIMIZATION ===============================================================

## Add the new constraint for two bags capacity
m.addConstrs(quicksum(weight[item] * isPacked[bag, item] for item in Items) 
                     <= bagCapacities[bag] for bag in Bags)

## C2: The item cannot be packed in multiple bags
PACKED = 1

### The sum of the binary var `item` (over all bags '*') cannot exceed a count of 1
for item in Items:
        m.addConstr(isPacked.sum('*', item) <= PACKED)

## Maximize the intrinsic value of the packed items for each bag
objFun = quicksum(quicksum(value[item] * isPacked[bag, item] 
                           for item in Items) 
                  for bag in Bags)

## Set objective, update the model, and Optimize
m.setObjective(objFun, GRB.MAXIMIZE)
m.update()
m.optimize()

print('\n\nPART II: MULTIPLE BAGS')
printOptimalSolution(m, isPacked, weight, value, Items, Bags)

















